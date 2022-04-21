# -*- coding: utf-8 -*-
"""
Created on Fri Apr 15 13:32:00 2022

@author: DELL
自动读取珠江新城建筑信息，计算并输出各结构非结构数量和种类
"""

import openpyxl
import xlrd
import numpy as np
#import win32com.client

# 读取结构层数和面积信息

xl = xlrd.open_workbook(r'D:\1HQY\Fragility\NCEstimate\ZJXCbuildingsAndInformation.xls')
sheet1 = xl.sheet_by_index(0)

StoryNumberlist = sheet1.col_values(0)
Arealist = sheet1.col_values(1)


# 读取整个工作簿的数据
data = openpyxl.load_workbook(filename="FEMAP-58_NormativeQuantityEstimationTool_042213.xlsm",
	                          keep_vba=True,
	                          read_only=False)

# 取到某个sheet的数据
table = data['Normative Quantity Estimate']

c = np.zeros((35,945))

for ii in range(584,946):
    Area = Arealist[ii-1]
    table.cell(10,5).value = Area
# 追加结束后保存
    data.save("FEMAP-58_NormativeQuantityEstimationTool_042213.xlsm")
    #xlvba = win32com.client.Dispatch("Excel.applicatioan")
    #xlvba.Workbooks.Open(Filename=r"D:\1HQY\Fragility\NCEstimate\FEMAP-58_NormativeQuantityEstimationTool_042213.xlsm")
    #xlvba.Application.Run('compile_fragility.MyVBA')
    
    from win32com.client import Dispatch
    app = Dispatch("Excel.Application")
    app.Visible = True
    xlbook=app.Workbooks.Open(r'D:\1HQY\Fragility\NCEstimate\FEMAP-58_NormativeQuantityEstimationTool_042213.xlsm') 
    app.Run("compile_fragility")  # 宏的名字
    xlbook.Close(SaveChanges=True)
    app.Quit()
    
    data = openpyxl.load_workbook(filename="FEMAP-58_NormativeQuantityEstimationTool_042213.xlsm",
    	                          keep_vba=True,
    	                          read_only=False)

    # 取到某个sheet的数据
    table = data['Normative Quantity Estimate']
    
    for kk in range(11,46):
       c[kk-11,ii-1] = table.cell(kk,24).value
    
    
